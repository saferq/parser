import requests
import re
from bs4 import BeautifulSoup
from pprint import pprint as pp
import openpyxl


class Scraper():
    """ Парсим сайт https://nd.gostinfo.ru """

    def create_soup(self, url):
        """ Готовим текст """
        res = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
        soup = BeautifulSoup(res.text, "lxml")
        return soup

    def get_table(self, soup):
        """ Получаем нужные строки талицы сайта """
        div_table = soup.select(
            '#cphCenter_cphCenter_ctl00_Panel_ND'
        )
        find_table = div_table[0].find('table').find_all('table')
        rows = find_table[1].find_all('tr', valign='top')
        return rows

    def get_table_test(self, soup):
        """ Получаем нужные строки талицы сайта """
        div_table = soup.select(
            '#cphCenter_cphCenter_ctl00_Panel_ND'
        )
        find_table = div_table[0].find('table').find_all('table')
        rows = find_table[1].find_all('tr', valign='top')
        return rows

    def iter_row(self, rows):
        """ Перебираем строки и сохраняем в список url """
        urls = []
        for row in rows:
            cell = row.find_all('td')
            # ------------------------------
            # Получение ссылки
            t_url = cell[0].select('div > a')
            for txt in t_url:
                t_url = 'https://nd.gostinfo.ru' + txt.get('href')
                urls.append(t_url)
           # ------------------------------
            # Тут обозначение
            # t_number = re.sub(r'\s+', ' ', cell[0].text)[1:]
            # Тут наименование
            # t_name = re.sub(r'\s+', ' ', cell[1].text)[1:]
            # print(t_name)
            # ------------------------------
            # Тут статус
            # t_status = re.sub(r'\s+', ' ', cell[2].text)
            # if t_status[0:1] == ' ':
            #     print(t_status[1:])
            # else:
            #     print(t_status)
            # print('')
        return urls

    # ------------------------------
    # Работа с листом
    def parsing_page(self, url, number_row, sheet):
        """ Получаем информацию со странице """
        res = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
        soup = BeautifulSoup(res.text, "lxml")

        txt = self.get_text_row(soup, 'Обозначение')
        print(txt + ' ' + url)
        sheet.cell(row=number_row+1, column=1, value=txt)
        txt = self.get_text_row(soup, 'Заглавие на русском языке')
        # print(txt)
        sheet.cell(row=number_row+1, column=2, value=txt)
        txt = self.get_text_row(soup, 'Статус')
        # print(txt)
        sheet.cell(row=number_row+1, column=3, value=txt)
        txt = self.get_text_row(soup, 'Обозначение заменяющего')
        sheet.cell(row=number_row+1, column=4, value=txt)
        txt = self.get_text_row(soup, 'Обозначение заменяемого(ых) ')
        sheet.cell(row=number_row+1, column=5, value=txt)

    #

    def get_text_row(self, soup, find_tag):
        n = soup.find('td', text=find_tag)
        if n == None:
            row = ''
            return row
        else:
            row = n.parent.find_all('td')
            return row[1].text

    # ------------------------------
    # Основная функция
    def get_main(self, url, total_pages, current_page, number_row):
        """ Главный цикл """
        file = 'ntd2.xlsx'
        wb = openpyxl.load_workbook(file)
        sheet = wb['NTD']
        while current_page <= total_pages:
            print(f'\nСтраница {str(current_page + 1)}')
            soup = self.create_soup(f'{url}{str(current_page)}')
            rows = self.get_table(soup)
            urls = self.iter_row(rows)
            for link in urls:
                self.parsing_page(link, number_row, sheet)
                number_row += 1
            wb.save(file)
            current_page += 1


if __name__ in '__main__':
    url = "https://nd.gostinfo.ru/doc.aspx?control=&search=&sort=%20ASC&catalogid=gost&classid=-1&s=-1&page="
    total_pages = 2312  # Индекс страниц
    current_page = 2197   # Индекс страницы
    number_row = 43817    # Индекс строки

    sc = Scraper()
    sc.get_main(url, total_pages, current_page, number_row)

    # current_page = 1
    # soup = sc.create_soup(f'{url}{current_page}')
    # tbl = sc.get_table_test(soup)
    # urls = sc.iter_row(tbl)
    # for url in urls:
    #     print(url)
